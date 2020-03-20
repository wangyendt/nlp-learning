#!/usr/bin/env python
# encoding: utf-8

"""
@author: Wayne
@contact: wangye.hope@gmail.com
@software: PyCharm
@file: vivo_comments
@time: 2019/10/12 14:51
"""

import os
import time
from urllib.parse import urlencode

import jsonpath
import pandas as pd
import requests
from fake_useragent import UserAgent
from openpyxl import load_workbook

products = {
    'nex3': 10001477,
    'iqoo': 10000467,
    'iqoo-pro': 10001399,
    'iqoo3': 10001922,
}

url_base = 'http://shop.vivo.com.cn/api/v1/remark/getDetail?'

ua = UserAgent()
last_id = ['']
to_search = 'iqoo3'
save_path = f'data/vivo-{to_search}-comments.txt'
save_excel_path = f'data/vivo-{to_search}-comments.xlsx'


def get_page_number():
    json_str = get_one_page_comments(
        url_base,
        to_search
    )
    page_numbers = jsonpath.jsonpath(json_str, '$..pages')
    return int(page_numbers[0])


def get_one_page_comments(url, product, page=1):
    header = {
        'user-agent': ua.random,
        # 'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.90 Safari/537.36'
    }
    data = {
        'spuId': products[product],
        'onlyHasPicture': False,
        'fullpaySkuIdSet': '',
        'pageNum': page,
        'pageSize': 10,
        't': int(time.time() * 1000),
        'lastId': last_id[0]
    }
    try:
        response = requests.get(url + urlencode(data), headers=header)
        time.sleep(4)
        if response.status_code == 200:
            return response.json()
        else:
            print(f'状态码错误{response.status_code}')
            return None
    except:
        print(f'爬取单页失败{url}')
        return None


def parse_one_page_comments(page):
    json_str = get_one_page_comments(
        url_base,
        'nex3', page
    )
    # import pprint
    # pprint.pprint(json_str)
    comments = jsonpath.jsonpath(json_str, '$..content')
    score = jsonpath.jsonpath(json_str, '$..summaryScore')
    last_id_tmp = jsonpath.jsonpath(json_str, '$..lastId')
    if last_id_tmp and last_id_tmp[0]: last_id[0] = last_id_tmp[0]
    print(f'page={page}, last_id: {last_id[0]}, last_id_tmp: {last_id_tmp}')
    if comments:
        for comm, scr in zip(comments, score):
            yield {'content': comm, 'score': int(scr)}


def clear_save_file():
    if os.path.exists(save_path):
        os.remove(save_path)
    if os.path.exists(save_excel_path):
        os.remove(save_excel_path)


def create_empty_file():
    nan_excel = pd.DataFrame()
    nan_excel.to_excel(save_excel_path)
    return pd.ExcelWriter(save_excel_path)


def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError

    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()


def save_to_file(contents, page):
    size = 100
    pd_, pr_ = divmod(page, size)
    sp = save_path.replace('.txt', f'_{pd_ * size + 1}-{(pd_ + 1) * size}.txt')
    sep = save_excel_path.replace('.xlsx', f'_{pd_ * size + 1}-{(pd_ + 1) * size}.xlsx')
    with open(sp, 'a+', encoding='utf-8') as f:
        for content in contents:
            # print(content)
            f.write(f'score: {content["score"]}, comment: {content["content"]}\n')
    append_df_to_excel(sep, pd.DataFrame(contents), header=None, index=False)
    # if os.path.exists(save_excel_path):
    #     old = pd.read_excel(save_excel_path, header=None)
    #     new = pd.DataFrame(contents)
    #     print(old.head(2))
    #     print(new.head(2))
    #     old.columns = ['1', '2']
    #     new.columns = ['1', '2']
    #     if not old:
    #         res = new.copy()
    #     else:
    #         res = old.append(new, ignore_index=True)
    # else:
    #     res = pd.DataFrame(contents)
    # res.to_excel(save_excel_path, header=None, index=None)


def work(page):
    print(f'正在爬取第{page}页...')
    return [dic for dic in parse_one_page_comments(page)]


if __name__ == '__main__':
    page_number = get_page_number()
    print(f'有{page_number}页需要爬取')
    clear_save_file()
    result = []
    for p in range(page_number):
        result += work(p)
        if result:
            print(result[-1])
        save_to_file(result, p)
    # pool = Pool()
    # max_worker = 4
    # n_tasks = page_number // max_worker
    # for i in range(n_tasks + 1):
    #     result = list(pool.imap(work, range(max_worker * i + 1, min(page_number, max_worker * (i + 1)) + 1)))
    #     print(f'截止目前共有{len(result)}条结果')
    #     # todo: 要改成监听异步减少运行时间
    #     save_to_file(sum(result, []))
    # writer.save()
